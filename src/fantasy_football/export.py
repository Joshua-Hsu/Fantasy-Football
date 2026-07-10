"""Draft-board export: a tiered cheat sheet as an Excel workbook.

One sheet per position plus an overall sheet. Within a position, players are
ordered by their user rating (the head-to-head result), grouped into tiers, with
both their rank-in-position and rank-in-tier shown alongside the auction value
and the three stat bases. Tiers reflect the user ratings, so the board sharpens
as you make picks.

openpyxl is an optional dependency (``pip install -e ".[export]"``).
"""

from __future__ import annotations

from typing import NamedTuple

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .models import UserRating
from .ratings import seed_ratings, user_rating_tiers
from .scoring import DEFAULT_RULES, ScoringRules
from .valuation import (
    ALL_POSITIONS,
    DEFAULT_LEAGUE,
    LeagueConfig,
    _latest_season,
    compute_values,
)

# A few soft fills cycled per tier so tier bands are easy to scan.
_TIER_FILLS = ["FFF6E5", "E8F0FE", "E9F7EF", "FCE8F3", "F3E8FD", "EAF2F8", "FFF0E6"]

_HEADERS = ["Tier", "PosRank", "TierRank", "Player", "Tm", "Ovr", "$", "UserRtg",
            "LastYr", "PPG", "3yrWtd"]


class BoardRow(NamedTuple):
    key: str
    tier: int
    pos_rank: int
    tier_rank: int
    name: str
    team: str
    overall_rank: int
    dollars: float
    user_rating: float
    total: float
    ppg: float
    w3yr: float
    position: str
    is_rookie: bool


def effective_pool_ratings(
    session: Session,
    rating_overrides: dict[str, float] | None,
    *,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
) -> tuple[dict[str, float], dict[str, int]]:
    """One rating + ONE tier numbering for the whole draftable pool.

    The master's continuous rating where a player has one (ladder-guarded, see
    write_tiers_csv), his production value otherwise — both on the same scale —
    then tiers derived from those ratings across the full pool. This is the fix
    for mixing two independently-numbered tier schemes (master vs value k-means),
    which let a backup missing from the master out-tier his own starter.
    """
    pre = compute_values(session, year=year, config=config, rules=rules, basis=basis)
    by_key = {r.key: r for rows in pre.values() for r in rows}
    overrides = rating_overrides or {}
    eff: dict[str, float] = {}
    for key, r in by_key.items():
        o = overrides.get(key)
        if o is not None and (o > 0.5 or r.basis_value <= 0.5):
            eff[key] = o
        else:
            eff[key] = r.basis_value
    return eff, derive_tiers_from_ratings(eff, by_key)


def build_board(
    session: Session,
    *,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
    manual_tiers: dict[str, int] | None = None,
    fixed_prices: dict[str, float] | None = None,
    rating_overrides: dict[str, float] | None = None,
) -> dict[str, list[BoardRow]]:
    """Per-position rows grouped by tier, with within-tier and overall ranks.

    With ``rating_overrides`` (the master's continuous ratings) the whole pool
    gets a single tier numbering via :func:`effective_pool_ratings` — players
    absent from the master slot in by production value, so a lower producer can
    never out-tier a higher one unless the user rated him higher. Without
    overrides, falls back to the legacy DB user-rating tiers merged with
    ``manual_tiers``. ``fixed_prices`` pins expected market prices for specific
    players; the field re-prices around them.
    """
    if rating_overrides:
        ratings, tiers = effective_pool_ratings(
            session, rating_overrides, year=year, config=config, rules=rules, basis=basis
        )
    else:
        seed_ratings(session, year=year, config=config, rules=rules, basis=basis)
        tiers = user_rating_tiers(session)
        if manual_tiers:
            tiers = {**tiers, **manual_tiers}
        ratings = {r.key: r.rating for r in session.scalars(select(UserRating))}
    values = compute_values(
        session, year=year, config=config, rules=rules, basis=basis, manual_tiers=tiers,
        fixed_prices=fixed_prices,
    )

    board: dict[str, list[BoardRow]] = {}
    for pos in ALL_POSITIONS:
        rows = values.get(pos, [])
        if not rows:
            continue
        # Group by tier (contiguous), best rating first within each tier.
        ranked = sorted(rows, key=lambda r: (r.tier, -ratings.get(r.key, r.basis_value)))
        per_tier_seen: dict[int, int] = {}
        out: list[BoardRow] = []
        for i, r in enumerate(ranked, 1):
            per_tier_seen[r.tier] = per_tier_seen.get(r.tier, 0) + 1
            out.append(
                BoardRow(
                    key=r.key, tier=r.tier, pos_rank=i, tier_rank=per_tier_seen[r.tier],
                    name=r.name, team=r.team, overall_rank=r.overall_rank,
                    dollars=r.dollars,
                    user_rating=round(ratings.get(r.key, r.basis_value), 1),
                    total=r.total, ppg=r.ppg, w3yr=r.w3yr, position=pos,
                    is_rookie=r.is_rookie,
                )
            )
        board[pos] = out
    return board


#: Default pick-game depth per position — roughly the draftable universe for a
#: 12-team league plus a buffer. Trims the long tail of players unlikely to be
#: drafted so you never have to weigh them head-to-head.
DEFAULT_WEBAPP_DEPTH = {"QB": 24, "RB": 60, "WR": 64, "TE": 28, "K": 16, "DST": 32}

#: Positions limited to one per team (NFL starters) rather than a depth cap.
#: QB/K backups don't get drafted, so we keep only each team's starter. DST is
#: left uncapped (32) so every defense is ranked.
STARTER_POSITIONS = {"QB", "K"}

#: Positions covered per-team instead of by a global cap, so every team's
#: starter(s) appear: RB top 2 (covers contested backfields), WR top 3.
PER_TEAM_DEPTH = {"RB": 2, "WR": 3}


def build_webapp_data(
    session: Session,
    *,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
    depth: int | dict[str, int] | None = None,
    rookie_max_round: int = 3,
    manual_tiers: dict[str, int] | None = None,
    seed_overrides: dict[str, float] | None = None,
    pinned_tiers: dict[str, int] | None = None,
    prices: dict[str, float] | None = None,
    backups: dict[str, tuple[str | None, str]] | None = None,
    backup_overrides: dict[str, str] | None = None,
) -> dict[str, list[dict]]:
    """Per-position player data for the static pick game (browser app).

    Each entity carries the three comparison stats, a seed value (for the
    in-browser Elo), current team, and coaching. No DB needed at play time.

    ``depth`` caps how many *veterans* per position are included (top N by value),
    so the long tail of undraftable players is left out. Incoming rookies drafted
    in rounds 1..``rookie_max_round`` are added on top (seeded by draft capital so
    they interleave with veterans), since they have no stats to rank them.
    ``manual_tiers`` (hand-set tiers, by key) keeps those players in the pool.
    ``seed_overrides`` (by key) is the master's continuous rating: when present a
    player starts the game from that rating instead of raw value, so the app
    refines the master rather than starting from scratch each week.
    ``prices`` (master Rec$ per key), ``backups`` (depth-chart gsis map) and
    ``backup_overrides`` (starter name -> backup name) feed the in-browser
    personal draft packet: each entity carries its price and most-likely backup.
    """
    from .models import Player, Team

    manual_tiers = manual_tiers or {}
    seed_overrides = seed_overrides or {}
    prices = prices or {}
    backups = backups or {}
    backup_overrides = backup_overrides or {}
    if isinstance(depth, int):
        caps = {pos: depth for pos in ALL_POSITIONS}
    else:
        caps = {**DEFAULT_WEBAPP_DEPTH, **(depth or {})}

    # With master ratings, tier/price the whole pool on ONE numbering (players
    # missing from the master slot in by value) — same fix as the packet.
    # manual_tiers itself stays as the master's keys: it also drives pool
    # retention below, which must not balloon to every player.
    tier_map = manual_tiers
    if seed_overrides:
        _eff, tier_map = effective_pool_ratings(
            session, seed_overrides, year=year, config=config, rules=rules, basis=basis
        )
    # Commissioner pins (the master's tier_pin column) override the derived
    # numbering — the admin board is law until released.
    if pinned_tiers:
        tier_map = dict(tier_map)
        tier_map.update(pinned_tiers)
    values = compute_values(
        session, year=year, config=config, rules=rules, basis=basis, manual_tiers=tier_map
    )
    coaching = {
        t.abbreviation: (t.head_coach or "", t.offensive_coordinator or "",
                         bool(t.hc_new), bool(t.oc_new))
        for t in session.scalars(select(Team))
    }
    draft = {
        f"p{slug}": (rnd, pick)
        for slug, rnd, pick in session.execute(
            select(Player.slug, Player.draft_round, Player.draft_pick).where(Player.draft_round.isnot(None))
        )
        if slug
    }
    # Last-year stat lines for the cards / CSV export.
    last_year = year or _latest_season(session)
    pstats = _player_last_year(session, last_year) if last_year else {}
    toff = _team_offense(session, last_year) if last_year else {}
    dstats = _dst_last_year(session, last_year) if last_year else {}
    ages = _player_ages(session, (last_year or 0) + 1) if last_year else {}
    byes = _team_byes(session)
    tshares = _target_shares(session, last_year) if last_year else {}
    rshares = _rush_shares(session, last_year) if last_year else {}
    totals = _fantasy_totals(session, last_year, rules) if last_year else []
    fppg = {t[0]: t[6] for t in totals}
    fname_ppg = {t[1].lower(): t[6] for t in totals}

    def backup_for(r, pool) -> tuple[str, object]:
        """(most-likely backup name, his fantasy PPG or "") — same resolution
        order as the packet: manual override, depth chart, same-team next."""
        override = backup_overrides.get(r.name.lower())
        if override:
            return override, fname_ppg.get(override.lower(), "")
        gsis = r.key[1:] if r.key.startswith("p") else None
        if gsis and gsis in backups:
            bk_gsis, bk_name = backups[gsis]
            return bk_name, fppg.get(f"p{bk_gsis}", "") if bk_gsis else ""
        seen = False
        for other in pool:  # value-sorted; the next same-team player after r
            if other.key == r.key:
                seen = True
            elif seen and other.team == r.team:
                return other.name, other.ppg
        return "", ""

    # Where a rookie of each round slots into a position's value scale.
    round_slot = {1: 7, 2: 17, 3: 27}

    def seed_for(r, fallback):
        # Seed from the master's continuous rating when we have one (so the app
        # refines last week's master); otherwise fall back to the player's value.
        # Either way the seed is a continuous value-scale number, not a tier band,
        # so the gap between players reflects a real difference and picks move
        # someone past their neighbours. Ladder ratings (<= 0.5, see
        # write_tiers_csv) are synthetic "unrated" markers — ignore them.
        if r.key in seed_overrides and (seed_overrides[r.key] > 0.5 or fallback <= 0.5):
            return seed_overrides[r.key]
        return fallback

    def emit(r, seed, pool):
        hc, oc, hc_new, oc_new = coaching.get(r.team, ("", "", False, False))
        rnd, pick = draft.get(r.key, (None, None))
        row = {
            "key": r.key, "name": r.name, "team": r.team, "pos": r.position,
            "total": r.total, "ppg": r.ppg, "w3yr": r.w3yr,
            "seed": round(seed, 1), "rookie": r.is_rookie, "hc": hc, "oc": oc,
            "stat": _format_statline(r.key, r.position, pstats, dstats),
            "tmoff": _format_team_context(r.team, r.position, toff),
            "cols": _stat_columns(r.key, r.position, r.team, pstats, toff, dstats,
                                  ages, byes, tshares, rshares),
        }
        # The master's tier for this player (pins already applied): seeds the
        # commissioner's #/admin board. The pick game itself stays rating-based.
        if tier_map.get(r.key):
            row["tier"] = tier_map[r.key]
        if hc_new:
            row["hcN"] = 1
        if oc_new:
            row["ocN"] = 1
        if r.is_rookie and rnd:
            row["draft"] = f"R{rnd} P{pick}"
        # Master market pin when present, else the computed (tier-monotonic)
        # auction value — so the personal packet always shows a Rec$.
        # Whole dollars: auction bids can't include cents.
        row["price"] = max(1, round(prices.get(r.key, r.dollars)))
        if r.position != "DST":
            bk_name, bk_ppg = backup_for(r, pool)
            if bk_name:
                row["bkp"] = bk_name
                row["bkp_ppg"] = bk_ppg
        return row

    def rookie_seed(r, vet_seeds):
        rnd, pick = draft.get(r.key, (None, None))
        idx = round_slot.get(rnd, 30)
        base = vet_seeds[min(idx, len(vet_seeds) - 1)] if vet_seeds else 0.0
        return base - (pick or 0) * 0.001  # earlier picks rank a touch higher

    out: dict[str, list[dict]] = {}
    for pos in ALL_POSITIONS:
        rows = values.get(pos, [])  # sorted by value desc
        vets = [r for r in rows if not r.is_rookie]
        vet_seeds = [r.basis_value for r in vets]
        rookies = [
            r for r in rows
            if r.is_rookie and (draft.get(r.key, (99,))[0] or 99) <= rookie_max_round
        ]

        if pos in STARTER_POSITIONS:
            # One per team (the projected starter): best seed per current team.
            best: dict[str, tuple] = {}
            for r in vets + rookies:
                base = r.basis_value if not r.is_rookie else rookie_seed(r, vet_seeds)
                seed = seed_for(r, base)
                team = r.team or r.key
                if team not in best or seed > best[team][1]:
                    best[team] = (r, seed)
            chosen = list(best.values())
        elif pos in PER_TEAM_DEPTH:
            # Union: keep the global depth pool (never drop anyone already in)
            # AND add top N per team so every team's starter(s) are covered.
            picked = list(vets[: caps.get(pos)])
            by_team: dict[str, list] = {}
            for r in vets:
                by_team.setdefault(r.team or r.key, []).append(r)
            for members in by_team.values():
                members.sort(key=lambda r: r.basis_value, reverse=True)
                picked += members[: PER_TEAM_DEPTH[pos]]
            seen: set = set()
            uniq: list = []
            for r in picked:
                if r.key not in seen:
                    seen.add(r.key)
                    uniq.append(r)
            chosen = [(r, seed_for(r, r.basis_value)) for r in uniq]
            chosen += [(r, seed_for(r, rookie_seed(r, vet_seeds))) for r in rookies]
        else:
            chosen = [(r, seed_for(r, r.basis_value)) for r in vets[: caps.get(pos)]]
            chosen += [(r, seed_for(r, rookie_seed(r, vet_seeds))) for r in rookies]

        # Always include any player you've explicitly tiered (manual_tiers),
        # even if they fall outside the per-team / depth selection.
        have = {r.key for r, _ in chosen}
        for r in rows:
            if r.key in manual_tiers and r.key not in have:
                chosen.append((r, seed_for(r, r.basis_value)))
                have.add(r.key)

        chosen.sort(key=lambda x: x[1], reverse=True)
        out[pos] = [emit(r, s, rows) for r, s in chosen]
    return out


def write_webapp_data(
    session: Session,
    path: str,
    *,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
    depth: int | dict[str, int] | None = None,
    manual_tiers: dict[str, int] | None = None,
    seed_overrides: dict[str, float] | None = None,
    pinned_tiers: dict[str, int] | None = None,
    prices: dict[str, float] | None = None,
    backups: dict[str, tuple[str | None, str]] | None = None,
    starters: dict[tuple[str, str], list[str]] | None = None,
    backup_overrides: dict[str, str] | None = None,
    leaders: dict[str, int] | None = None,
    base: str | None = None,
) -> str:
    """Write the pick-game data as ``docs/data.js`` (``window.FF_DATA = {...}``).

    Besides the per-position pools, the payload carries everything the
    in-browser personal draft packet needs: per-entity prices/backups (see
    :func:`build_webapp_data`), a ``teams`` table (coaching, PF, offense totals,
    skill depth chart) and a ``top200`` box-stats table.
    """
    import json

    from .models import Team

    data = build_webapp_data(
        session, year=year, config=config, rules=rules, basis=basis, depth=depth,
        manual_tiers=manual_tiers, seed_overrides=seed_overrides,
        pinned_tiers=pinned_tiers,
        prices=prices, backups=backups, backup_overrides=backup_overrides,
    )
    starters = starters or {}
    last_year = year or _latest_season(session)
    totals = _fantasy_totals(session, last_year, rules) if last_year else []
    pstats = _player_last_year(session, last_year) if last_year else {}
    toff = _team_offense(session, last_year) if last_year else {}
    pf = _team_points_for(session, last_year) if last_year else {}

    tdef = _team_defense_totals(session, last_year) if last_year else {}
    tvol = _team_volume(session, last_year) if last_year else {}
    vacated = _vacated_shares(session, last_year) if last_year else {}

    teams_payload = []
    for t in sorted(session.scalars(select(Team)), key=lambda t: -pf.get(t.abbreviation, 0)):
        abbr = t.abbreviation
        off = toff.get(abbr, {})
        if not off and abbr not in pf:
            continue
        yds, plays = off.get("total_yards", 0), off.get("plays", 0)
        d = tdef.get(abbr, {})
        games = d.get("games", 0)
        vol = tvol.get(abbr, {})
        vac_t, vac_r = vacated.get(abbr, ("", ""))
        wrs = starters.get((abbr, "WR"), [])
        rbs = starters.get((abbr, "RB"), [])
        teams_payload.append({
            "team": abbr, "hc": t.head_coach or "", "oc": t.offensive_coordinator or "",
            "hcN": 1 if t.hc_new else 0, "ocN": 1 if t.oc_new else 0,
            "pf": pf.get(abbr, ""), "pa": d.get("pa", ""),
            "pag": round(d["pa"] / games, 1) if games else "",
            "yds": yds or "", "ydsg": round(yds / games, 1) if games and yds else "",
            "plays": plays or "", "ypp": round(yds / plays, 1) if plays else "",
            "pass": off.get("pass", ""), "passAtt": vol.get("pass_att", ""),
            "passRk": off.get("pass_rank", ""),
            "rush": off.get("rush", ""), "rushAtt": vol.get("rush_att", ""),
            "rushRk": off.get("rush_rank", ""),
            "td": (vol.get("pass_td", 0) + vol.get("rush_td", 0)) or "",
            "patd": vol.get("pass_td", ""), "rutd": vol.get("rush_td", ""),
            "vacTgt": vac_t, "vacRush": vac_r,
            "qb": ", ".join(starters.get((abbr, "QB"), [])[:1]),
            "rb": rbs[0] if len(rbs) > 0 else "",
            "rb2": rbs[1] if len(rbs) > 1 else "",
            "wr1": wrs[0] if len(wrs) > 0 else "", "wr2": wrs[1] if len(wrs) > 1 else "",
            "wr3": wrs[2] if len(wrs) > 2 else "",
            "te": ", ".join(starters.get((abbr, "TE"), [])[:1]),
        })

    top200 = []
    skill = [t for t in totals if t[2] not in ("K", "QB")]  # QBs get their own table
    for key, name, pos, team, g, fpts, ppg in skill[:200]:
        s = pstats.get(key, {})
        top200.append([
            name, team, pos, g, fpts, ppg,
            s.get("rush_attempts", 0), s.get("rush_yards", 0), s.get("rush_touchdowns", 0),
            s.get("targets", 0), s.get("receptions", 0),
            s.get("receiving_yards", 0), s.get("receiving_touchdowns", 0),
        ])
    qbstats = []
    for key, name, pos, team, g, fpts, ppg in [t for t in totals if t[2] == "QB"][:40]:
        s = pstats.get(key, {})
        qbstats.append([
            name, team, g, fpts, ppg,
            s.get("pass_yards", 0), s.get("pass_touchdowns", 0),
            s.get("interceptions_thrown", 0),
            s.get("rush_attempts", 0), s.get("rush_yards", 0), s.get("rush_touchdowns", 0),
        ])

    # Leaderboard for the levels page: keep it bounded (top 1000 committers).
    top_leaders = dict(sorted((leaders or {}).items(),
                              key=lambda kv: -kv[1])[:1000])
    payload = {
        "basis": basis, "positions": data, "stat_headers": CSV_STAT_HEADERS,
        "leaders": top_leaders,
        "teams": teams_payload,
        "year": last_year,
        # Which master this build came from: commits echo it, and the rebuild
        # only blends pick files made against the master it is rebuilding.
        "base": base or "",
        "top200_headers": ["Player", "Tm", "Pos", "G", "FPTS", "PPG",
                           "RuAtt", "RuYds", "RuTD",
                           "Tgt", "Rec", "ReYds", "ReTD"],
        "top200": top200,
        "qb_headers": ["Player", "Tm", "G", "FPTS", "PPG",
                       "PaYds", "PaTD", "INT", "RuAtt", "RuYds", "RuTD"],
        "qbstats": qbstats,
    }
    with open(path, "w") as fh:
        fh.write("// Generated by `fantasy_football build-webapp` - do not edit by hand.\n")
        fh.write("window.FF_DATA = ")
        json.dump(payload, fh, ensure_ascii=False)
        fh.write(";\n")
    return path


# Last-year stat columns shown inline per position: (header, PlayerGameStats attr).
# RB: rushing + full receiving line (RBs catch). WR/TE: receiving only (no rush).
_RECEIVING = [
    ("Tgt", "targets"), ("Rec", "receptions"), ("ReYds", "receiving_yards"),
    ("ReTD", "receiving_touchdowns"), ("Tgt%", "__tgtshare"), ("RZTgt", "redzone_targets"),
]
_RB_STATS = [("Car", "rush_attempts"), ("RuYds", "rush_yards"), ("RuTD", "rush_touchdowns")] + _RECEIVING
STAT_COLS: dict[str, list[tuple[str, str]]] = {
    "QB": [("PaAtt", "pass_attempts"), ("PaYds", "pass_yards"), ("PaTD", "pass_touchdowns"),
           ("INT", "interceptions_thrown"), ("RuAtt", "rush_attempts"), ("RuYds", "rush_yards")],
    "RB": _RB_STATS, "WR": _RECEIVING, "TE": _RECEIVING,
    "K": [("FGM", "field_goals_made"), ("FGA", "field_goals_attempted"), ("XPM", "extra_points_made")],
    "DST": [("PA/g", "PA"), ("Sack", "Sack"), ("INT", "INT"), ("DefTD", "TD")],
}
#: Positions that show their team's offensive context inline.
TEAM_OFFENSE_POSITIONS = {"QB", "RB", "WR", "TE"}


def _player_last_year(session: Session, year: int) -> dict[str, dict[str, int]]:
    """{p<id>: {stat: 2025 regular-season total}} for all the stats we display."""
    from .models import Game, Player, PlayerGameStats

    # Only real PlayerGameStats columns (skip DST/target-share sentinels).
    cols = sorted({attr for cols in STAT_COLS.values() for _, attr in cols
                   if hasattr(PlayerGameStats, attr)})
    aggs = [func.sum(getattr(PlayerGameStats, c)).label(c) for c in cols]
    query = (
        select(Player.slug, *aggs)
        .join(PlayerGameStats, PlayerGameStats.player_id == Player.id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Player.slug)
    )
    out: dict[str, dict[str, int]] = {}
    for row in session.execute(query):
        if row[0]:
            out[f"p{row[0]}"] = {c: int(v or 0) for c, v in zip(cols, row[1:])}
    return out


def _team_offense(session: Session, year: int) -> dict[str, dict[str, int]]:
    """{team_abbr: {total_yards, plays, rush, pass, *_rank}} for the season."""
    from .models import Game, Team, TeamGameStats

    query = (
        select(
            Team.abbreviation,
            func.sum(TeamGameStats.total_yards), func.sum(TeamGameStats.plays),
            func.sum(TeamGameStats.rushing_yards), func.sum(TeamGameStats.passing_yards),
        )
        .join(TeamGameStats, TeamGameStats.team_id == Team.id)
        .join(Game, Game.id == TeamGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Team.id)
    )
    data: dict[str, dict[str, int]] = {}
    for abbr, ty, pl, ru, pa in session.execute(query):
        data[abbr] = {"total_yards": int(ty or 0), "plays": int(pl or 0),
                      "rush": int(ru or 0), "pass": int(pa or 0)}
    for key in ("total_yards", "plays", "rush", "pass"):
        for rank, abbr in enumerate(sorted(data, key=lambda t: data[t][key], reverse=True), 1):
            data[abbr][key + "_rank"] = rank
    return data


def _dst_last_year(session: Session, year: int) -> dict[str, dict[str, float]]:
    """{d<abbr>: {PA (per game), Sack, INT, TD}} from team defensive stats."""
    from .models import Game, Team, TeamGameStats

    query = (
        select(
            Team.abbreviation, func.count(TeamGameStats.id),
            func.sum(TeamGameStats.points_allowed), func.sum(TeamGameStats.sacks),
            func.sum(TeamGameStats.interceptions),
            func.sum(TeamGameStats.defensive_tds), func.sum(TeamGameStats.special_teams_tds),
        )
        .join(TeamGameStats, TeamGameStats.team_id == Team.id)
        .join(Game, Game.id == TeamGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Team.id)
    )
    out: dict[str, dict[str, float]] = {}
    for abbr, g, pa, sk, inte, dtd, sttd in session.execute(query):
        g = int(g or 0)
        out[f"d{abbr}"] = {
            "PA": round((pa or 0) / g, 1) if g else 0.0, "Sack": int(sk or 0),
            "INT": int(inte or 0), "TD": int((dtd or 0) + (sttd or 0)),
        }
    return out


def _format_statline(key: str, pos: str, pstats: dict, dstats: dict) -> str:
    """Compact last-year stat line for a player/defense (position-aware)."""
    if pos == "DST":
        d = dstats.get(key, {})
        return f"{d.get('PA', 0)}PA/g {d.get('Sack', 0)}sk {d.get('INT', 0)}int {d.get('TD', 0)}td"
    s = pstats.get(key)
    if not s:
        return ""
    if pos == "QB":
        return (f"{s['pass_yards']}yd {s['pass_touchdowns']}td {s['interceptions_thrown']}int, "
                f"{s['rush_attempts']}car {s['rush_yards']}ryd")
    if pos == "K":
        return f"{s['field_goals_made']}/{s['field_goals_attempted']}FG {s['extra_points_made']}XP"
    return (f"{s['receptions']}rec {s['receiving_yards']}yd {s['receiving_touchdowns']}td, "
            f"{s['rush_attempts']}car {s['rush_yards']}ryd {s['rush_touchdowns']}rtd")


def _format_team_context(team: str, pos: str, toff: dict) -> str:
    """Team-offense rank summary for offensive players (blank otherwise)."""
    if pos not in TEAM_OFFENSE_POSITIONS:
        return ""
    o = toff.get(team)
    if not o:
        return ""
    return f"off yds#{o['total_yards_rank']} pass#{o['pass_rank']} rush#{o['rush_rank']}"


# Canonical individual stat columns (each sortable) used by the CSVs and app.
CSV_STAT_HEADERS = [
    "Bye", "Age",
    "PaYds", "PaTD", "INT", "RuAtt", "RuYds", "RuTD", "Tgt", "Rec", "ReYds", "ReTD",
    "Tgt%", "Rush%", "RZTgt",
    "FGM", "FGA", "XPM", "DefPA", "DefSk", "DefINT", "DefTD",
    "TmYds", "TmYdsRk", "TmPlays", "TmPlaysRk", "TmRush", "TmRushRk", "TmPass", "TmPassRk",
]


def _player_ages(session: Session, year: int) -> dict[str, int]:
    """{p<id>: age as of Sept 1 of the season year} from birth dates."""
    import datetime as dt

    from .models import Player

    ref = dt.date(year, 9, 1)
    out: dict[str, int] = {}
    for slug, bd in session.execute(select(Player.slug, Player.birth_date)):
        if bd and slug:
            out[f"p{slug}"] = (ref - bd).days // 365
    return out


def _team_byes(session: Session) -> dict[str, int]:
    """{team_abbr: bye_week}."""
    from .models import Team

    return {t.abbreviation: t.bye_week for t in session.scalars(select(Team)) if t.bye_week}


def _target_shares(session: Session, year: int) -> dict[str, float]:
    """{p<id>: average per-game target share %}.

    For each game a player appeared in, target share = his targets / his team's
    targets that game; we average those across the games he played (so missed
    games don't dilute it).
    """
    from collections import defaultdict

    from .models import Game, Player, PlayerGameStats

    rows = session.execute(
        select(Player.slug, PlayerGameStats.game_id,
               PlayerGameStats.team_id, PlayerGameStats.targets)
        .join(PlayerGameStats, PlayerGameStats.player_id == Player.id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
    ).all()

    team_game_targets: dict[tuple, int] = defaultdict(int)
    for _slug, gid, tid, tg in rows:
        team_game_targets[(tid, gid)] += int(tg or 0)

    per_game: dict[str, list[float]] = defaultdict(list)
    for slug, gid, tid, tg in rows:
        tt = team_game_targets[(tid, gid)]
        if tt > 0 and slug:
            per_game[slug].append(int(tg or 0) / tt)

    return {
        f"p{slug}": round(sum(shares) / len(shares) * 100, 1)
        for slug, shares in per_game.items() if shares
    }


def _stat_columns(key: str, pos: str, team: str, pstats: dict, toff: dict, dstats: dict,
                  ages: dict | None = None, byes: dict | None = None,
                  tshares: dict | None = None, rshares: dict | None = None) -> dict:
    """One value per CSV_STAT_HEADERS column for an entity (blank where N/A)."""
    c: dict[str, object] = {h: "" for h in CSV_STAT_HEADERS}
    c["Age"] = (ages or {}).get(key, "")
    c["Bye"] = (byes or {}).get(team, "")
    c["Tgt%"] = (tshares or {}).get(key, "") if pos in ("RB", "WR", "TE") else ""
    c["Rush%"] = (rshares or {}).get(key, "") if pos in ("RB", "WR", "TE") else ""
    if pos == "DST":
        d = dstats.get(key, {})
        c["DefPA"], c["DefSk"] = d.get("PA", ""), d.get("Sack", "")
        c["DefINT"], c["DefTD"] = d.get("INT", ""), d.get("TD", "")
        return c
    s = pstats.get(key)
    if s:
        if pos == "QB":
            c["PaYds"], c["PaTD"], c["INT"] = s["pass_yards"], s["pass_touchdowns"], s["interceptions_thrown"]
            c["RuAtt"], c["RuYds"], c["RuTD"] = s["rush_attempts"], s["rush_yards"], s["rush_touchdowns"]
        elif pos == "K":
            c["FGM"], c["FGA"], c["XPM"] = s["field_goals_made"], s["field_goals_attempted"], s["extra_points_made"]
        else:  # RB / WR / TE: receiving (+ rushing for RB only)
            c["Tgt"], c["Rec"] = s["targets"], s["receptions"]
            c["ReYds"], c["ReTD"] = s["receiving_yards"], s["receiving_touchdowns"]
            c["RZTgt"] = s.get("redzone_targets", "")
            if pos == "RB":
                c["RuAtt"], c["RuYds"], c["RuTD"] = s["rush_attempts"], s["rush_yards"], s["rush_touchdowns"]
    if pos in TEAM_OFFENSE_POSITIONS:
        o = toff.get(team)
        if o:
            c["TmYds"], c["TmYdsRk"] = o["total_yards"], o["total_yards_rank"]
            c["TmPlays"], c["TmPlaysRk"] = o["plays"], o["plays_rank"]
            c["TmRush"], c["TmRushRk"] = o["rush"], o["rush_rank"]
            c["TmPass"], c["TmPassRk"] = o["pass"], o["pass_rank"]
    return c


def _safe_cell(value):
    """Neutralize spreadsheet formula injection in a text cell.

    A CSV value beginning with =, +, -, @ (or a control char) is executed as a
    formula by Excel/Sheets. Prefix such values with a single quote so they're
    treated as text. Numbers are returned unchanged.
    """
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def derive_tiers_from_ratings(
    ratings: dict[str, float], by_key: dict
) -> dict[str, int]:
    """Derive integer tiers per position from a continuous rating map.

    Groups keys by position, ranks them best->worst by rating, and applies the
    same gap-aware sizing the rest of the toolkit uses, so the master's tiers
    reflect the continuous rating (which the pick game refines).
    """
    from .valuation import DEFAULT_TIER_K, assign_sized_tiers

    out: dict[str, int] = {}
    by_pos: dict[str, list[str]] = {}
    for key in ratings:
        r = by_key.get(key)
        pos = r.position if r else None
        if pos:
            by_pos.setdefault(pos, []).append(key)
    for pos, keys in by_pos.items():
        keys.sort(key=lambda k: ratings[k], reverse=True)
        out.update(assign_sized_tiers(
            keys, [ratings[k] for k in keys], DEFAULT_TIER_K.get(pos, 6)
        ))
    return out


#: Confidence half-point for the user-rating blend: this many head-to-head
#: comparisons of a player gives the league's rating a 50% say vs his anchor.
CONFIDENCE_PICKS = 6


def write_tiers_csv(
    session: Session,
    path: str,
    *,
    tiers: dict[str, int] | None = None,
    ratings: dict[str, float] | None = None,
    user_ratings: dict[str, float] | None = None,
    comps: dict[str, int] | None = None,
    pinned_ratings: dict[str, float] | None = None,
    pinned_tiers: dict[str, int] | None = None,
    carried_tiers: dict[str, int] | None = None,
    confidence: int = CONFIDENCE_PICKS,
    prices: dict[str, float] | None = None,
    notes: dict[tuple[str, int], str] | None = None,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
) -> str:
    """Write an enriched master/tiers CSV (rating + tiers + stats + prices).

    ``ratings`` is the anchor map (the previous master's continuous ratings);
    any player without one defaults to his production value, so the master
    always carries a rating for the whole pool. ``user_ratings`` is this
    round's merged pick-game ratings, blended onto the anchors with a
    confidence weight ``w = comps / (comps + confidence)`` — a couple of picks
    nudge a player, many picks dominate. Pick files with no comps information
    (legacy) get full weight, which is the old override behavior. Pass
    ``tiers`` directly for the legacy integer-only path. ``notes`` is the
    hand-written tier description per (position, tier) — it's repeated on each
    row of that tier so the CSV round-trips it (and you can edit it in place).

    ``pinned_tiers`` is the commissioner's explicit tier per key from a FRESH
    #/admin overwrite — literal law, applied after the gap derivation.
    ``carried_tiers`` are pins carried forward from the previous master's
    ``tier_pin`` column: they act as rating BANDS anchored on the previous
    geometry, so crowd picks that move a blended rating across a band
    boundary promote/demote the player to the neighbouring tier. (The CLI
    feeds the blend only pick files made against this same master, so only
    same-base picks can cross admin boundaries.) Both kinds are written back
    to ``tier_pin`` — pinned positions stay pinned until released.
    """
    import csv

    prices = prices or {}
    last_year = year or _latest_season(session)
    values = compute_values(session, year=year, config=config, rules=rules, basis=basis)
    by_key = {r.key: r for rows in values.values() for r in rows}

    if ratings is not None or user_ratings or pinned_ratings:
        ratings = ratings or {}
        user_ratings = user_ratings or {}
        comps = comps or {}
        pinned_ratings = pinned_ratings or {}
        # Default a rating for every selected player (value when not picked), so
        # the master seeds the whole app; then derive tiers from those ratings.
        # Repair pass: early masters carried a synthetic near-zero "ladder"
        # (0, -0.1, -0.2, ...) for players nobody had rated. Real ratings live on
        # the value scale (hundreds), so anything <= 0.5 for a player with actual
        # production is unrated — reseed it from value instead of letting the
        # ladder wreck the tier derivation (phantom gaps -> singleton tiers).
        keys = (set(ratings) | set(user_ratings) | set(pinned_ratings)
                | set(pinned_tiers or {}) | set(tiers or {}))
        anchors = {
            k: ratings.get(k, by_key[k].basis_value)
            for k in keys if k in by_key
        }
        anchors = {
            k: (by_key[k].basis_value
                if r <= 0.5 and by_key[k].basis_value > 0.5 else r)
            for k, r in anchors.items()
        }
        # Confidence blend: the league's pick-game rating pulls a player away
        # from his anchor in proportion to how much he was actually compared.
        ratings = dict(anchors)
        for k, ur in user_ratings.items():
            if k not in by_key:
                continue
            if ur <= 0.5 and by_key[k].basis_value > 0.5:
                continue  # ladder value riding along in a refresh: not a real pick
            # Unknown comps (legacy/foreign submissions) count as a single
            # comparison - low influence, never full override, since the
            # public site accepts submissions from anyone.
            n = max(comps.get(k, 0), 1)
            w = n / (n + confidence)
            ratings[k] = round(w * ur + (1 - w) * anchors.get(k, by_key[k].basis_value), 2)
        # Admin pins win outright: the commissioner's drag-and-drop overwrite
        # is applied after the crowd blend, no confidence discount.
        for k, pr in (pinned_ratings or {}).items():
            if k in by_key:
                ratings[k] = round(pr, 2)
        tiers = derive_tiers_from_ratings(ratings, by_key)
        # Commissioner tier pins: full manual override of the derived tiers.
        # A fresh admin overwrite (pinned_tiers) is literal. Carried pins
        # (carried_tiers) act as bands on the previous master's geometry:
        # every player in the position is re-assigned by where his BLENDED
        # rating falls between the anchor boundaries, so fresh crowd picks
        # with enough weight move him across an admin boundary. Positions
        # containing a pin are then renumbered contiguous 1..K in
        # (tier, rating) order, so pins can't leave gaps or inverted labels.
        pinned_tiers = {k: t for k, t in (pinned_tiers or {}).items() if k in tiers}
        carried_tiers = {k: t for k, t in (carried_tiers or {}).items() if k in tiers}
        admin_pos = {by_key[k].position for k in pinned_tiers}
        band_pos = {by_key[k].position for k in carried_tiers} - admin_pos
        if pinned_tiers or carried_tiers:
            for pos in band_pos:
                carried_in = [k for k in carried_tiers if by_key[k].position == pos]
                carried_in.sort(key=lambda k: (carried_tiers[k], -anchors.get(k, 0.0)))
                bounds = [
                    (anchors.get(carried_in[i], 0.0)
                     + anchors.get(carried_in[i + 1], 0.0)) / 2
                    for i in range(len(carried_in) - 1)
                    if carried_tiers[carried_in[i]] != carried_tiers[carried_in[i + 1]]
                ]
                for k in [k2 for k2 in tiers if by_key[k2].position == pos]:
                    r = ratings.get(k, 0.0)
                    tiers[k] = 1 + sum(1 for b in bounds if r < b)
            for k, t in pinned_tiers.items():
                tiers[k] = t
            pinned_pos = admin_pos | band_pos
            for pos in pinned_pos:
                keys_in = [k for k in tiers if by_key[k].position == pos]
                keys_in.sort(key=lambda k: (tiers[k], -ratings.get(k, 0.0)))
                next_label, last_seen = 0, None
                for k in keys_in:
                    if tiers[k] != last_seen:
                        next_label, last_seen = next_label + 1, tiers[k]
                    tiers[k] = next_label
                # Redistribute the continuous ratings to FIT the pinned tiers:
                # tight spacing within a tier, a clear gap at every boundary.
                # Board order is preserved and the position keeps its top
                # rating, so pick-game seeds, the next crowd blend, and a
                # later "release" all see the structure the admin drew (the
                # gap comfortably clears the 5%-of-spread derivation floor).
                vals = [ratings.get(k, 0.0) for k in keys_in]
                top = max(vals) if vals else 0.0
                span = top - min(vals) if vals else 0.0
                step, gap = 2.0, max(8.0, 0.06 * span)
                bounds = (tiers[keys_in[-1]] - 1) if keys_in else 0
                within = max(len(keys_in) - 1 - bounds, 0)
                need = within * step + bounds * gap
                if need > 0 and top > 1 and need > top - 1:
                    scale = (top - 1) / need  # never redistribute below ~1
                    step, gap = step * scale, gap * scale
                cur, prev_t = top, None
                for k in keys_in:
                    if prev_t is not None:
                        cur -= gap if tiers[k] != prev_t else step
                    prev_t = tiers[k]
                    ratings[k] = round(cur, 2)
            # Record the FINAL labels for EVERY player in a pinned position
            # (new arrivals join the structure) so the carried-forward pins
            # match what this master actually says.
            pinned_tiers = {k: tiers[k] for k in tiers
                            if by_key[k].position in pinned_pos}
    else:
        tiers = tiers or {}
        ratings = {}
        pinned_tiers = {}
    pstats = _player_last_year(session, last_year) if last_year else {}
    toff = _team_offense(session, last_year) if last_year else {}
    dstats = _dst_last_year(session, last_year) if last_year else {}
    ages = _player_ages(session, (last_year or 0) + 1) if last_year else {}
    byes = _team_byes(session)
    tshares = _target_shares(session, last_year) if last_year else {}
    rshares = _rush_shares(session, last_year) if last_year else {}

    notes = notes or {}
    with open(path, "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["key", "manual_tier", "rating", "tier_pin", "tier_note", "name",
                         "pos", "team", "total", "ppg"] + CSV_STAT_HEADERS + ["price"])
        # Order by tier then rating so the file reads top-to-bottom.
        ordered = sorted(tiers, key=lambda k: (tiers[k], -ratings.get(k, 0.0)))
        for key in ordered:
            tier = tiers[key]
            r = by_key.get(key)
            pos = r.position if r else ""
            cols = _stat_columns(key, pos, r.team if r else "", pstats, toff, dstats,
                                 ages, byes, tshares, rshares)
            rating = ratings.get(key)
            writer.writerow(
                [_safe_cell(key), tier, round(rating, 2) if rating is not None else "",
                 pinned_tiers.get(key, ""),
                 _safe_cell(notes.get((pos, tier), "")),
                 _safe_cell(r.name if r else key), _safe_cell(pos),
                 _safe_cell(r.team if r else ""), r.total if r else "", r.ppg if r else ""]
                + [cols[h] for h in CSV_STAT_HEADERS]
                + [max(1, round(prices[key])) if key in prices else ""]
            )
    return path


def _fantasy_totals(session: Session, year: int, rules: ScoringRules):
    """All players' fantasy production for a season, best first.

    Returns ``[(key, name, pos, team, games, fpts, ppg), ...]`` — powers both
    the Top-200 sheet and backup-PPG lookups in the packet.
    """
    from .models import Game, Player, PlayerGameStats
    from .scoring import score_expression

    points = func.sum(score_expression(rules))
    games = func.count(PlayerGameStats.id)
    query = (
        select(Player.slug, Player.full_name, Player.position, Player.current_team,
               games, points)
        .join(PlayerGameStats, PlayerGameStats.player_id == Player.id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Player.id)
        .order_by(points.desc())
    )
    out = []
    for slug, name, pos, team, g, total in session.execute(query):
        if not slug:
            continue
        g, total = int(g or 0), float(total or 0.0)
        ppg = max(total / g, 0.0) if g else 0.0  # display floor: no negative PPG
        out.append((f"p{slug}", name, pos or "", team or "",
                    g, round(total, 1), round(ppg, 1)))
    return out


def _team_points_for(session: Session, year: int) -> dict[str, int]:
    """{team_abbr: regular-season points scored} from game results."""
    from .models import Game, Team

    pf: dict[str, int] = {}
    teams = {t.id: t.abbreviation for t in session.scalars(select(Team))}
    games = session.scalars(
        select(Game).where(Game.season_year == year, Game.season_type == "regular")
    )
    for g in games:
        if g.home_score is None or g.away_score is None:
            continue
        home, away = teams.get(g.home_team_id), teams.get(g.away_team_id)
        if home:
            pf[home] = pf.get(home, 0) + g.home_score
        if away:
            pf[away] = pf.get(away, 0) + g.away_score
    return pf


def _team_defense_totals(session: Session, year: int) -> dict[str, dict[str, int]]:
    """{team_abbr: {pa, games}} — points allowed + games played, from results."""
    from .models import Game, Team

    teams = {t.id: t.abbreviation for t in session.scalars(select(Team))}
    out: dict[str, dict[str, int]] = {}
    games = session.scalars(
        select(Game).where(Game.season_year == year, Game.season_type == "regular")
    )
    for g in games:
        if g.home_score is None or g.away_score is None:
            continue
        for team_id, allowed in ((g.home_team_id, g.away_score), (g.away_team_id, g.home_score)):
            abbr = teams.get(team_id)
            if abbr:
                d = out.setdefault(abbr, {"pa": 0, "games": 0})
                d["pa"] += allowed
                d["games"] += 1
    return out


def _team_volume(session: Session, year: int) -> dict[str, dict[str, int]]:
    """{team_abbr: {pass_att, rush_att, pass_td, rush_td}} from player lines.

    TeamGameStats doesn't carry attempts or TD splits, so aggregate the player
    box scores by the team each line was recorded for.
    """
    from .models import Game, PlayerGameStats, Team

    query = (
        select(
            Team.abbreviation,
            func.sum(PlayerGameStats.pass_attempts), func.sum(PlayerGameStats.rush_attempts),
            func.sum(PlayerGameStats.pass_touchdowns), func.sum(PlayerGameStats.rush_touchdowns),
        )
        .join(PlayerGameStats, PlayerGameStats.team_id == Team.id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Team.id)
    )
    return {
        abbr: {"pass_att": int(pa or 0), "rush_att": int(ra or 0),
               "pass_td": int(ptd or 0), "rush_td": int(rtd or 0)}
        for abbr, pa, ra, ptd, rtd in session.execute(query)
    }


def _vacated_shares(session: Session, year: int) -> dict[str, tuple[float, float]]:
    """{team_abbr: (vacated target %, vacated rush-attempt %)}.

    The share of last season's targets / rush attempts that belonged to players
    who are no longer on that team (moved or inactive) — the volume available
    for rookies and new arrivals to slot into.
    """
    from .models import Game, Player, PlayerGameStats, Team

    rows = session.execute(
        select(Team.abbreviation, Player.current_team, Player.active,
               func.sum(PlayerGameStats.targets), func.sum(PlayerGameStats.rush_attempts))
        .join(PlayerGameStats, PlayerGameStats.team_id == Team.id)
        .join(Player, Player.id == PlayerGameStats.player_id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
        .group_by(Team.id, Player.id)
    ).all()

    totals: dict[str, list[float]] = {}
    gone: dict[str, list[float]] = {}
    for abbr, cur_team, active, tgt, ratt in rows:
        tgt, ratt = int(tgt or 0), int(ratt or 0)
        totals.setdefault(abbr, [0, 0])
        totals[abbr][0] += tgt
        totals[abbr][1] += ratt
        if cur_team != abbr or not active:
            gone.setdefault(abbr, [0, 0])
            gone[abbr][0] += tgt
            gone[abbr][1] += ratt
    out: dict[str, tuple[float, float]] = {}
    for abbr, (t_tgt, t_ratt) in totals.items():
        g_tgt, g_ratt = gone.get(abbr, [0, 0])
        out[abbr] = (
            round(g_tgt / t_tgt * 100, 1) if t_tgt else 0.0,
            round(g_ratt / t_ratt * 100, 1) if t_ratt else 0.0,
        )
    return out


def _rush_shares(session: Session, year: int) -> dict[str, float]:
    """{p<id>: average per-game share of his team's rush attempts %}.

    Mirrors :func:`_target_shares` for the ground game.
    """
    from collections import defaultdict

    from .models import Game, Player, PlayerGameStats

    rows = session.execute(
        select(Player.slug, PlayerGameStats.game_id,
               PlayerGameStats.team_id, PlayerGameStats.rush_attempts)
        .join(PlayerGameStats, PlayerGameStats.player_id == Player.id)
        .join(Game, Game.id == PlayerGameStats.game_id)
        .where(Game.season_year == year, Game.season_type == "regular")
    ).all()

    team_game_att: dict[tuple, int] = defaultdict(int)
    for _slug, gid, tid, att in rows:
        team_game_att[(tid, gid)] += int(att or 0)

    per_game: dict[str, list[float]] = defaultdict(list)
    for slug, gid, tid, att in rows:
        ta = team_game_att[(tid, gid)]
        if ta > 0 and slug:
            per_game[slug].append(int(att or 0) / ta)

    return {
        f"p{slug}": round(sum(shares) / len(shares) * 100, 1)
        for slug, shares in per_game.items() if shares
    }


# Packet position tabs mirror the hand-made packet layout; bid columns are
# left blank to write in during the auction (they feed the Draft Board).


def write_cheatsheet(
    session: Session,
    path: str,
    *,
    year: int | None = None,
    config: LeagueConfig = DEFAULT_LEAGUE,
    rules: ScoringRules = DEFAULT_RULES,
    basis: str = "w3yr",
    manual_tiers: dict[str, int] | None = None,
    fixed_prices: dict[str, float] | None = None,
    tier_notes: dict[tuple[str, int], str] | None = None,
    backups: dict[str, tuple[str | None, str]] | None = None,
    starters: dict[tuple[str, str], list[str]] | None = None,
    backup_overrides: dict[str, str] | None = None,
    ratings: dict[str, float] | None = None,
) -> str:
    """Write the draft packet to an .xlsx file. Returns the path.

    Sheets: a live **Draft Board** (mark a player drafted + enter the price and
    remaining recommendations re-adjust for auction inflation), one tab per
    position laid out in tier sections (tier note | team | PPG | starter | bid |
    backup PPG | most-likely backup | bid), a **Team Stats** tab (coaching,
    offense totals, skill depth chart), and a **Top 200** box-stats tab.

    ``tier_notes`` is {(pos, tier): text} — your hand-written tier descriptions
    (auto "$ range" labels where missing). ``backups`` is {gsis_id:
    (backup_gsis, backup_name)} from the depth charts; ``backup_overrides`` is
    {starter name lowercased: backup name} and wins; with neither, the backup
    falls back to the next same-team player at the position in the board.
    ``starters`` is {(team, pos): [names]} for the Team Stats depth columns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    board = build_board(
        session, year=year, config=config, rules=rules, basis=basis,
        manual_tiers=manual_tiers, fixed_prices=fixed_prices,
        rating_overrides=ratings,
    )
    tier_notes = tier_notes or {}
    backups = backups or {}
    starters = starters or {}
    backup_overrides = backup_overrides or {}

    last_year = year or _latest_season(session)
    totals = _fantasy_totals(session, last_year, rules) if last_year else []
    fppg = {t[0]: t[6] for t in totals}          # {p<gsis>: fantasy PPG}
    fname_ppg = {t[1].lower(): t[6] for t in totals}  # by name, for override rows
    player_stats = _player_last_year(session, last_year) if last_year else {}
    team_off = _team_offense(session, last_year) if last_year else {}
    pf = _team_points_for(session, last_year) if last_year else {}

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    wb.remove(wb.active)

    def _tier_fill(tier: int) -> PatternFill:
        return PatternFill("solid", fgColor=_TIER_FILLS[(tier - 1) % len(_TIER_FILLS)])

    def _backup_for(r: BoardRow, pool: list[BoardRow]):
        """(backup name, fantasy PPG or "", entity key or None) for a row."""
        override = backup_overrides.get(r.name.lower())
        if override:
            return override, fname_ppg.get(override.lower(), ""), None
        gsis = r.key[1:] if r.key.startswith("p") else None
        if gsis and gsis in backups:
            bk_gsis, bk_name = backups[gsis]
            return (bk_name, fppg.get(f"p{bk_gsis}", "") if bk_gsis else "",
                    f"p{bk_gsis}" if bk_gsis else None)
        # Fallback: next same-team player at this position in the board.
        for other in pool:
            if other.team == r.team and other.pos_rank > r.pos_rank:
                return other.name, other.ppg, other.key
        return "", "", None

    tshares = _target_shares(session, last_year) if last_year else {}
    rshares = _rush_shares(session, last_year) if last_year else {}

    # --- Per-position tier tabs --------------------------------------------
    # Volume-share columns per position: how much of the team's passing /
    # rushing volume the player owns (the "how safe is his role" signal).
    share_cols = {"RB": ["Tgt%", "Rush%"], "WR": ["Tgt%", "Rush%"],
                  "TE": ["Tgt%", "Rush%"]}
    # Where each player's Bid cell lives, so the Draft Board can watch it.
    bid_cells: dict[str, tuple[str, str]] = {}  # key -> (sheet title, cell ref)

    def _position_sheet(pos: str, rows: list[BoardRow]) -> None:
        ws = wb.create_sheet(pos[:31])
        has_backup = pos != "DST"
        shares = share_cols.get(pos, [])
        headers = (["Tier", "Team", "PPG", "Starter"] + shares + ["Rec$", "Bid"]
                   + (["Bkp PPG", "Backup"]
                      + ["Bkp " + sh for sh in shares]
                      + ["Bkp Bid"] if has_backup else []))
        bid_col = headers.index("Bid") + 1
        recd_col = headers.index("Rec$") + 1
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = center

        current_tier = None
        for r in rows:
            if current_tier is not None and r.tier != current_tier:
                ws.append([])  # visual gap between tiers
            first_of_tier = r.tier != current_tier
            current_tier = r.tier
            if first_of_tier:
                group = [x for x in rows if x.tier == r.tier]
                dollars = [x.dollars for x in group]
                label = tier_notes.get((pos, r.tier)) or (
                    f"Tier {r.tier} — ${round(min(dollars))}-{round(max(dollars))}"
                    if dollars else f"Tier {r.tier}"
                )
            name = r.name + (" (R)" if r.is_rookie else "")
            ppg = "" if r.is_rookie else r.ppg
            line: list = [label if first_of_tier else "", r.team, ppg, name]
            for sh in shares:
                src = tshares if sh == "Tgt%" else rshares
                line.append(src.get(r.key, ""))
            line += [r.dollars, None]
            if has_backup:
                bk_name, bk_ppg, bk_key = _backup_for(r, rows)
                line += [bk_ppg, bk_name]
                for sh in shares:
                    src_map = tshares if sh == "Tgt%" else rshares
                    line.append(src_map.get(bk_key, "") if bk_key else "")
                line.append(None)
            ws.append(line)
            row_i = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_i, column=c).fill = _tier_fill(r.tier)
            ws.cell(row=row_i, column=1).alignment = wrap
            ws.cell(row=row_i, column=recd_col).number_format = '"$"0'
            bid_cells[r.key] = (ws.title, f"{get_column_letter(bid_col)}{row_i}")

        ws.freeze_panes = "A2"
        widths = ([26, 6, 7, 22] + [6] * len(shares) + [7, 7]
                  + ([8, 20] + [7] * len(shares) + [8] if has_backup else []))
        for c, w in zip(range(1, len(headers) + 1), widths):
            ws.column_dimensions[get_column_letter(c)].width = w

    for pos in ALL_POSITIONS:
        if pos in board:
            _position_sheet(pos, board[pos])

    # --- Team Stats tab ------------------------------------------------------
    from .models import Team

    tdef = _team_defense_totals(session, last_year) if last_year else {}
    tvol = _team_volume(session, last_year) if last_year else {}
    vacated = _vacated_shares(session, last_year) if last_year else {}

    ws = wb.create_sheet("Team Stats")
    ts_headers = ["Rk", "Team", "HC", "OC", "PF", "PA", "PA/G",
                  "Yds", "Yds/G", "Plays", "Y/P",
                  "PassYds", "PassAtt", "PassRk", "RushYds", "RushAtt", "RushRk",
                  "TD", "PaTD", "RuTD", "VacTgt%", "VacRush%",
                  "QB", "RB1", "RB2", "WR1", "WR2", "WR3", "TE"]
    ws.append(ts_headers)
    for c in range(1, len(ts_headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = center
    teams = sorted(session.scalars(select(Team)), key=lambda t: -pf.get(t.abbreviation, 0))
    rank = 0
    for t in teams:
        abbr = t.abbreviation
        off = team_off.get(abbr, {})
        if not off and abbr not in pf:
            continue  # inactive franchises
        rank += 1
        yds, plays = off.get("total_yards", 0), off.get("plays", 0)
        d = tdef.get(abbr, {})
        games = d.get("games", 0)
        vol = tvol.get(abbr, {})
        vac_t, vac_r = vacated.get(abbr, ("", ""))
        wrs = starters.get((abbr, "WR"), [])
        rbs = starters.get((abbr, "RB"), [])
        ws.append([
            rank, abbr, t.head_coach or "TBD", t.offensive_coordinator or "TBD",
            pf.get(abbr, ""), d.get("pa", ""),
            round(d["pa"] / games, 1) if games else "",
            yds or "", round(yds / games, 1) if games and yds else "",
            plays or "", round(yds / plays, 1) if plays else "",
            off.get("pass", ""), vol.get("pass_att", ""), off.get("pass_rank", ""),
            off.get("rush", ""), vol.get("rush_att", ""), off.get("rush_rank", ""),
            (vol.get("pass_td", 0) + vol.get("rush_td", 0)) or "",
            vol.get("pass_td", ""), vol.get("rush_td", ""),
            vac_t, vac_r,
            ", ".join(starters.get((abbr, "QB"), [])[:1]),
            rbs[0] if len(rbs) > 0 else "", rbs[1] if len(rbs) > 1 else "",
            wrs[0] if len(wrs) > 0 else "", wrs[1] if len(wrs) > 1 else "",
            wrs[2] if len(wrs) > 2 else "",
            ", ".join(starters.get((abbr, "TE"), [])[:1]),
        ])
        # New-to-role coaches get a soft accent shade — worth noting on draft day.
        new_fill = PatternFill("solid", fgColor="E9FAEC")
        if t.hc_new:
            ws.cell(row=ws.max_row, column=3).fill = new_fill
        if t.oc_new:
            ws.cell(row=ws.max_row, column=4).fill = new_fill
    ws.freeze_panes = "C2"
    for c, w in zip(range(1, len(ts_headers) + 1),
                    (4, 6, 17, 17, 6, 6, 6, 8, 7, 7, 5, 8, 8, 7, 8, 8, 7,
                     5, 6, 6, 8, 8, 15, 15, 15, 15, 15, 15, 15)):
        ws.column_dimensions[get_column_letter(c)].width = w

    # --- Stats tabs: skill-player Top 200 + a dedicated QB tab --------------
    # (no kickers anywhere; QBs get their own sheet with passing-first columns)
    ws = wb.create_sheet(f"{last_year} Top 200 Stats"[:31])
    t2_headers = ["Rk", "Player", "Tm", "Pos", "G", "FPTS", "PPG",
                  "RuAtt", "RuYds", "RuTD",
                  "Tgt", "Rec", "ReYds", "ReTD"]
    ws.append(t2_headers)
    for c in range(1, len(t2_headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = center
    skill = [t for t in totals if t[2] not in ("K", "QB")]
    for i, (key, name, pos, team, g, fpts, ppg) in enumerate(skill[:200], 1):
        s = player_stats.get(key, {})
        ws.append([
            i, name, team, pos, g, fpts, max(ppg, 0),
            s.get("rush_attempts", 0), s.get("rush_yards", 0), s.get("rush_touchdowns", 0),
            s.get("targets", 0), s.get("receptions", 0),
            s.get("receiving_yards", 0), s.get("receiving_touchdowns", 0),
        ])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    ws.column_dimensions["B"].width = 22
    for c in ("A", "C", "D", "E", "F", "G"):
        ws.column_dimensions[c].width = 7

    ws = wb.create_sheet(f"{last_year} QB Stats"[:31])
    qb_headers = ["Rk", "Player", "Tm", "G", "FPTS", "PPG",
                  "PaYds", "PaTD", "INT", "RuAtt", "RuYds", "RuTD"]
    ws.append(qb_headers)
    for c in range(1, len(qb_headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = center
    qbs = [t for t in totals if t[2] == "QB"]
    for i, (key, name, pos, team, g, fpts, ppg) in enumerate(qbs[:40], 1):
        s = player_stats.get(key, {})
        ws.append([
            i, name, team, g, fpts, max(ppg, 0),
            s.get("pass_yards", 0), s.get("pass_touchdowns", 0),
            s.get("interceptions_thrown", 0),
            s.get("rush_attempts", 0), s.get("rush_yards", 0), s.get("rush_touchdowns", 0),
        ])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    ws.column_dimensions["B"].width = 22
    for c in ("A", "C", "D", "E", "F"):
        ws.column_dimensions[c].width = 7

    # --- Live Draft Board: recommended prices that react to picks ----------
    _draft_sheet(wb, board, config, header_font, center, _tier_fill, bid_cells)
    wb.move_sheet("Draft Board", -(len(wb.sheetnames) - 1))  # make it first

    wb.save(path)
    return path


# Draft Board column layout (1-indexed):
#  A Pos  B Tier  C Player  D Base$  E Rec$  F Drafted  G Paid  H Weight(hidden)
#  I UserRtg  J LastYr  K PPG  L 3yr  M Tm  N Ovr  O PosBid ; controls in P/Q.
_DRAFT_HEADERS = ["Pos", "Tier", "Player", "Base$", "Rec$", "Drafted", "Paid",
                  "Weight", "UserRtg", "LastYr", "PPG", "3yrWtd", "Tm", "Ovr",
                  "PosBid"]


def _draft_sheet(wb, board, config, header_font, center, tier_fill,
                 bid_cells: dict[str, tuple[str, str]] | None = None) -> None:
    ws = wb.create_sheet("Draft Board")
    bid_cells = bid_cells or {}
    ws.append(_DRAFT_HEADERS)
    for c in range(1, len(_DRAFT_HEADERS) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = center

    rows = sorted(
        (r for rows in board.values() for r in rows),
        key=lambda r: r.dollars, reverse=True,
    )
    last = len(rows) + 1  # last data row (header is row 1)

    drafted = f"$F$2:$F${last}"
    paid = f"$G$2:$G${last}"
    weight = f"$H$2:$H${last}"

    for i, r in enumerate(rows, start=2):
        name = r.name + (" (R)" if r.is_rookie else "")
        ws.append([
            r.position, r.tier, name, r.dollars, None, None, None, None,
            r.user_rating, r.total, r.ppg, r.w3yr, r.team, r.overall_rank, None,
        ])
        # PosBid mirrors the player's Bid cell on his position tab, so a bid
        # written there flows straight into the board.
        sheet_cell = bid_cells.get(r.key)
        if sheet_cell:
            sheet, cell = sheet_cell
            ws.cell(row=i, column=15).value = f"='{sheet}'!{cell}"
        # Paid defaults to the position-tab bid; Drafted auto-marks once paid.
        # Both are plain formulas, so typing a value over them still works.
        ws.cell(row=i, column=7).value = f'=IF(O{i}<>"",O{i},"")'
        ws.cell(row=i, column=6).value = f'=IF(G{i}<>"","x","")'
        # Weight = max(Base$ - 1, 0); recompute defensively in-sheet.
        ws.cell(row=i, column=8).value = f"=MAX(D{i}-1,0)"
        # Rec$ = paid if drafted, else inflation-adjusted allocation of the
        # remaining pool across remaining weights (control block in column Q).
        ws.cell(row=i, column=5).value = (
            f'=IF(F{i}="x",G{i},'
            f'IF($Q$7>0,ROUND(1+H{i}/$Q$7*($Q$5-$Q$6),0),D{i}))'
        )
        for c in range(1, len(_DRAFT_HEADERS) + 1):
            ws.cell(row=i, column=c).fill = tier_fill(r.tier)

    # Control block (labels in P, live values in Q).
    controls = [
        ("Total pool", config.pool),
        ("Total slots", config.teams * config.roster_size),
        ("Spent", f'=SUMIFS({paid},{drafted},"x")'),
        ("Drafted", f'=COUNTIF({drafted},"x")'),
        ("Remaining pool", "=Q1-Q3"),
        ("Remaining slots", "=Q2-Q4"),
        ("Remaining weight", f'=SUMIFS({weight},{drafted},"<>x")'),
    ]
    for idx, (label, value) in enumerate(controls, start=1):
        ws.cell(row=idx, column=16, value=label).font = header_font
        ws.cell(row=idx, column=17, value=value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{last}"
    ws.column_dimensions["C"].width = 22
    for col in ("A", "B", "D", "E", "F", "G", "I", "J", "K", "L", "M", "N", "O"):
        ws.column_dimensions[col].width = 9
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["P"].width = 16
    for row in range(2, last + 1):
        for col in (4, 5, 7):  # Base$, Rec$, Paid
            ws.cell(row=row, column=col).number_format = '"$"0'
    ws.cell(row=1, column=17).number_format = '"$"0'  # total pool
    ws.cell(row=5, column=17).number_format = '"$"0'  # remaining pool
