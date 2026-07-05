"""Tests for the Excel draft-board export."""

from __future__ import annotations

import datetime as dt

import pytest

from fantasy_football.db import create_db_engine, get_sessionmaker, init_db
from fantasy_football.export import build_board, write_cheatsheet
from fantasy_football.models import Game, Player, PlayerGameStats, Season, Team
from fantasy_football.valuation import LeagueConfig


@pytest.fixture()
def session():
    engine = create_db_engine(":memory:")
    init_db(engine)
    with get_sessionmaker(engine)() as s:
        yield s


def _seed(session):
    gb = Team(abbreviation="GB", name="Packers")
    chi = Team(abbreviation="CHI", name="Bears")
    session.add_all([gb, chi, Season(year=2025)])
    session.flush()
    game = Game(season_year=2025, week=1, season_type="regular",
                game_date=dt.date(2025, 9, 7),
                home_team_id=gb.id, away_team_id=chi.id, home_score=20, away_score=10)
    session.add(game)
    session.flush()
    for i, yds in enumerate([150, 120, 90, 60, 40, 20]):
        p = Player(full_name=f"RB{i}", position="RB", slug=f"rb{i}")
        session.add(p)
        session.flush()
        session.add(PlayerGameStats(player_id=p.id, game_id=game.id, team_id=gb.id,
                                    rush_yards=yds, rush_attempts=yds // 10))
    k = Player(full_name="K0", position="K", slug="k0")
    session.add(k)
    session.flush()
    session.add(PlayerGameStats(player_id=k.id, game_id=game.id, team_id=gb.id,
                                fg_made_40_49=3, extra_points_made=2))
    session.commit()


def test_build_board_ranks_within_position_and_tier(session):
    _seed(session)
    board = build_board(session, year=2025, config=LeagueConfig(teams=1))
    rbs = board["RB"]
    assert [r.pos_rank for r in rbs] == list(range(1, len(rbs) + 1))
    # Tier ranks restart at 1 within each tier.
    first_in_tier = [r for r in rbs if r.tier_rank == 1]
    assert {r.tier for r in first_in_tier} == set(r.tier for r in rbs)
    assert rbs[0].name == "RB0"  # best RB on top


def test_write_cheatsheet_has_draft_board_and_live_formulas(session, tmp_path):
    _seed(session)
    out = tmp_path / "board.xlsx"
    write_cheatsheet(session, str(out), year=2025, config=LeagueConfig(teams=1))

    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert wb.sheetnames[0] == "Draft Board"   # live sheet first
    assert "RB" in wb.sheetnames               # static position sheet present

    ws = wb["Draft Board"]
    assert [c.value for c in ws[1][:7]] == ["Pos", "Tier", "Player", "Base$", "Rec$", "Drafted", "Paid"]
    # Rec$ (column E) is a live formula reacting to Drafted/Paid.
    assert str(ws["E2"].value).startswith("=IF(F2")
    # Control block (column P / index 16) carries the remaining-pool math.
    labels = [ws.cell(row=r, column=16).value for r in range(1, 8)]
    assert "Remaining pool" in labels and "Remaining weight" in labels


def test_csv_readers_reject_malicious_input(tmp_path):
    from fantasy_football.cli import _read_fixed_prices, _read_manual_tiers

    p = tmp_path / "evil.csv"
    p.write_text(
        "key,manual_tier,price\n"
        "=cmd|' /C calc'!A1,1,5\n"   # formula key -> rejected
        "p100,99,5\n"               # tier clamped to 30
        "p101,notanint,5\n"         # bad tier -> skipped
        "d ABC,1,5\n"               # space in key -> rejected
        "p102,2,=HYPERLINK(1)\n"    # bad price -> skipped
    )
    tiers = _read_manual_tiers(str(p))
    prices = _read_fixed_prices(str(p))
    assert all(k.startswith(("p", "d")) and "=" not in k and " " not in k for k in tiers)
    assert tiers["p100"] == 30            # clamped
    assert "p101" not in tiers            # bad tier skipped
    assert tiers["p102"] == 2
    assert "p102" not in prices           # bad price skipped


def test_merge_ratings_averages_across_files(tmp_path):
    from fantasy_football.cli import _merge_ratings

    a = tmp_path / "a.csv"
    a.write_text("key,rating\nprb0,100\nprb1,50\n")
    b = tmp_path / "b.csv"
    b.write_text("key,rating\nprb0,200\n")  # prb1 absent here
    merged = _merge_ratings([str(a), str(b)], base={"prb1": 999.0, "prb5": 10.0})
    assert merged["prb0"] == 150.0       # (100 + 200) / 2
    assert merged["prb1"] == 50.0        # only one file had it -> that value (not base)
    assert merged["prb5"] == 10.0        # nobody picked -> base carried forward


def test_master_round_trip_rating_and_derived_tiers(session, tmp_path):
    _seed(session)
    from fantasy_football.cli import _read_ratings
    from fantasy_football.export import write_tiers_csv

    out = tmp_path / "master.csv"
    # Hand the best RB a boosted rating; the writer should derive tiers from it.
    ratings = {"prb0": 500.0, "prb1": 120.0, "prb2": 90.0}
    write_tiers_csv(session, str(out), ratings=ratings, prices={"prb0": 42.0}, year=2025,
                    config=LeagueConfig(teams=1))
    text = out.read_text()
    assert text.splitlines()[0].startswith("key,manual_tier,rating,tier_note,name")
    back = _read_ratings(str(out))
    assert back["prb0"] == 500.0
    # Boosted prb0 lands in the top tier.
    import csv as _csv
    rows = {r["key"]: r for r in _csv.DictReader(out.open())}
    assert int(rows["prb0"]["manual_tier"]) == 1
    assert rows["prb0"]["price"] == "42.0"


def test_packet_position_sheet_layout_and_tabs(session, tmp_path):
    """Position tabs use the packet layout; Team Stats and Top 200 exist."""
    _seed(session)
    out = tmp_path / "packet.xlsx"
    write_cheatsheet(
        session, str(out), year=2025, config=LeagueConfig(teams=1),
        tier_notes={("RB", 1): "Bell cows - pay up"},
        backups={"rb0": ("rb5", "RB5")},          # depth chart: RB5 backs up RB0
        starters={("GB", "RB"): ["RB0"], ("GB", "QB"): ["Some QB"]},
        backup_overrides={"rb1": "Hand Picked"},  # manual fix wins for RB1
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    for tab in ("Draft Board", "RB", "Team Stats", "Top 200"):
        assert tab in wb.sheetnames

    ws = wb["RB"]
    assert [c.value for c in ws[1][:11]] == [
        "Tier", "Team", "PPG", "Starter", "Tgt%", "Rush%", "Rec$", "Bid",
        "Bkp PPG", "Backup", "Bkp Bid"]
    # Row 2 = best RB: hand-written tier note + depth-chart backup.
    assert ws["A2"].value == "Bell cows - pay up"
    assert ws["D2"].value == "RB0"
    assert float(ws["F2"].value) > 0        # RB0's rush-attempt share populated
    assert ws["J2"].value == "RB5"
    # RB1's backup comes from the manual override.
    rows = {ws.cell(row=r, column=4).value: r for r in range(2, ws.max_row + 1)}
    assert ws.cell(row=rows["RB1"], column=10).value == "Hand Picked"
    # A later RB with no depth entry falls back to next same-team RB.
    assert ws.cell(row=rows["RB2"], column=10).value == "RB3"

    ts = wb["Team Stats"]
    header = [c.value for c in ts[1]]
    assert header[:7] == ["Rk", "Team", "HC", "OC", "PF", "PA", "PA/G"]
    qb_col = header.index("QB") + 1
    rb2_col = header.index("RB2") + 1
    gb_row = next(r for r in range(2, ts.max_row + 1) if ts.cell(row=r, column=2).value == "GB")
    assert ts.cell(row=gb_row, column=5).value == 20            # PF from game score
    assert ts.cell(row=gb_row, column=6).value == 10            # PA from game score
    assert ts.cell(row=gb_row, column=qb_col).value == "Some QB"  # depth-chart starter

    t2 = wb["Top 200"]
    assert [c.value for c in t2[1][:7]] == ["Rk", "Player", "Tm", "Pos", "G", "FPTS", "PPG"]
    assert t2["B2"].value == "RB0"        # top scorer first
    names = [t2.cell(row=r, column=2).value for r in range(2, t2.max_row + 1)]
    assert "K0" not in names              # kickers excluded

    db = wb["Draft Board"]
    dh = [c.value for c in db[1]]
    assert dh[14] == "PosBid"     # column O; P/Q hold the control block
    # RB0 is the top row; his PosBid pulls the RB tab's Bid cell, Paid follows
    # PosBid, Drafted auto-marks from Paid.
    db_rows = {db.cell(row=r, column=3).value: r for r in range(2, db.max_row + 1)}
    r0 = db_rows["RB0"]
    assert str(db.cell(row=r0, column=15).value) == "='RB'!H2"
    assert str(db.cell(row=r0, column=7).value) == f'=IF(O{r0}<>"",O{r0},"")'
    assert str(db.cell(row=r0, column=6).value) == f'=IF(G{r0}<>"","x","")'


def test_depth_chart_mapping(monkeypatch):
    """Slot-aware backups: LWR2 backs up LWR1, not the other starting WR."""
    import pandas as pd

    from fantasy_football.ingest import nflverse as nv

    df = pd.DataFrame({
        "club_code": ["GNB"] * 4 + ["CHI"] * 2,
        "week": [3, 3, 3, 3, 3, 3],
        "depth_team": [1, 2, 1, 2, 1, 2],
        "position": ["WR", "WR", "WR", "WR", "QB", "QB"],
        "depth_position": ["LWR", "LWR", "RWR", "RWR", "QB", "QB"],
        "gsis_id": ["w1", "w2", "w3", "w4", "q1", "q2"],
        "full_name": ["Lwr One", "Lwr Two", "Rwr One", "Rwr Two", "Qb One", "Qb Two"],
    })
    monkeypatch.setattr(nv, "_fetch_depth_charts", lambda year: df)

    backups = nv.depth_backups(2025)
    assert backups["w1"] == ("w2", "Lwr Two")   # same slot, not the RWR
    assert backups["w3"] == ("w4", "Rwr Two")
    assert backups["q1"] == ("q2", "Qb Two")
    assert "w2" not in backups                   # deepest player has no backup

    starters = nv.depth_starters(2025)
    # Rank-1s first (GNB normalized to GB), then the rank-2s behind them.
    assert starters[("GB", "WR")][:2] == ["Lwr One", "Rwr One"]
    assert starters[("CHI", "QB")][0] == "Qb One"


def test_depth_chart_mapping_2026_schema(monkeypatch):
    """The 2026 nflverse depth files renamed every column; parser still works."""
    import pandas as pd

    from fantasy_football.ingest import nflverse as nv

    df = pd.DataFrame({
        "dt": ["2026-06-20"] * 2 + ["2026-07-01"] * 4,
        "team": ["GB", "GB", "GB", "GB", "CHI", "CHI"],
        "player_name": ["Old One", "Old Two", "Lwr One", "Lwr Two", "Qb One", "Qb Two"],
        "espn_id": [1, 2, 3, 4, 5, 6],
        "gsis_id": ["o1", "o2", "w1", "w2", "q1", "q2"],
        "pos_grp": ["offense"] * 6,
        "pos_abb": ["WR", "WR", "WR", "WR", "QB", "QB"],
        "pos_slot": ["LWR", "LWR", "LWR", "LWR", "QB", "QB"],
        "pos_rank": [1, 2, 1, 2, 1, 2],
    })
    monkeypatch.setattr(nv, "_fetch_depth_charts", lambda year: df)

    backups = nv.depth_backups(2026)
    assert backups["w1"] == ("w2", "Lwr Two")   # newest chart (dt) only
    assert "o1" not in backups                   # stale chart rows dropped
    assert nv.depth_starters(2026)[("CHI", "QB")][0] == "Qb One"


def test_monotonic_prices_follow_manual_tiers(session):
    """A lower manual tier can never carry a higher price than the tier above."""
    from fantasy_football.valuation import compute_values

    _seed(session)
    # Invert the tiers vs production: worst producers get the best tiers.
    manual = {"prb5": 1, "prb4": 1, "prb0": 2, "prb1": 2, "prb2": 3, "prb3": 3}
    values = compute_values(session, year=2025, config=LeagueConfig(teams=1),
                            manual_tiers=manual)
    rbs = values["RB"]
    by_tier: dict = {}
    for r in rbs:
        by_tier.setdefault(r.tier, []).append(r.dollars)
    tiers = sorted(by_tier)
    for hi, lo in zip(tiers, tiers[1:]):
        assert min(by_tier[hi]) >= max(by_tier[lo])  # monotonic down the board


def test_small_gap_between_near_equals_does_not_split_tier():
    """An 8.6-point dip on a 312-point scale (Gibbs vs Bijan) shares a tier."""
    from fantasy_football.valuation import assign_sized_tiers

    # Real shape of the RB master: two near-equals up top, a real cliff, then a
    # long smooth tail whose tiny gaps would otherwise drag the threshold down.
    values = [312.5, 303.9, 283.7, 271.7, 261.5, 261.2, 255.9, 250.3, 247.2]
    values += [float(v) for v in range(228, 80, -4)]  # smooth 4-point-gap tail
    keys = [f"k{i}" for i in range(len(values))]
    tiers = assign_sized_tiers(keys, values, 8)
    assert tiers["k0"] == tiers["k1"]            # Gibbs + Bijan together
    assert tiers["k1"] != tiers["k2"]            # the 20-point cliff still breaks


def test_unified_tiers_unmastered_player_cannot_outrank(session):
    """A player missing from the master slots in by value - his tier can never
    beat a mastered player who out-rates him (the Kyren/Corum inversion)."""
    from fantasy_football.export import build_board

    _seed(session)
    # Master rated only the top three RBs; rb3-rb5 are absent from it.
    overrides = {"prb0": 300.0, "prb1": 280.0, "prb2": 250.0}
    board = build_board(session, year=2025, config=LeagueConfig(teams=1),
                        rating_overrides=overrides)
    rbs = {r.name: r for r in board["RB"]}
    # Unmastered rb3 (lower production) must not out-tier mastered rb2.
    assert rbs["RB3"].tier >= rbs["RB2"].tier
    # And the board order follows production/rating order throughout.
    ordered = [r.name for r in board["RB"]]
    assert ordered == [f"RB{i}" for i in range(6)]


def test_qb_price_cap_and_redistribution(session):
    """Elite QB dollars clamp at the market cap; the excess flows to the field."""
    from fantasy_football.valuation import compute_values

    _seed(session)
    # An elite QB whose production dwarfs the RBs → raw VOR would price him huge.
    game = session.query(Game).first()
    qb = Player(full_name="Elite QB", position="QB", slug="qb0")
    qb2 = Player(full_name="Backup QB", position="QB", slug="qb1")
    session.add_all([qb, qb2])
    session.flush()
    session.add(PlayerGameStats(player_id=qb.id, game_id=game.id, team_id=game.home_team_id,
                                pass_yards=5000, pass_touchdowns=45, rush_yards=500))
    session.add(PlayerGameStats(player_id=qb2.id, game_id=game.id, team_id=game.home_team_id,
                                pass_yards=100))
    session.commit()

    capped = compute_values(session, year=2025, config=LeagueConfig(teams=1))
    uncapped = compute_values(session, year=2025,
                              config=LeagueConfig(teams=1, price_caps={}))
    qb_capped = capped["QB"][0]
    qb_raw = uncapped["QB"][0]
    assert qb_raw.dollars > 30.0            # raw VOR really would overprice him
    assert qb_capped.dollars == 30.0        # default market cap applied
    # The excess went to the uncapped positions (top RB got richer).
    assert capped["RB"][0].dollars > uncapped["RB"][0].dollars

    # POS=0 removes the cap via config.
    assert uncapped["QB"][0].dollars == qb_raw.dollars


def test_price_cap_cli_parsing():
    """--price-cap POS=N merges over the default; POS=0 removes it."""
    import argparse

    from fantasy_football.cli import _league_config

    args = argparse.Namespace(teams=12, budget=200, price_caps=["TE=20", "QB=0"])
    config = _league_config(args)
    assert config.price_caps == {"TE": 20.0}
    default = _league_config(argparse.Namespace(teams=12, budget=200, price_caps=None))
    assert default.price_caps == {"QB": 30.0}


def test_negative_ppg_clamped_to_zero():
    """A negative per-game average (bad DSTs, fumble-only lines) displays as 0."""
    from fantasy_football.valuation import _summarize

    ent = {"seasons": {2025: (10, -42.0)}}
    _summarize(ent, 2025)
    assert ent["ppg"] == 0.0
    assert ent["total"] == -42.0   # the true total is still visible


def test_ladder_ratings_reseed_from_value(session, tmp_path):
    """Synthetic near-zero 'ladder' ratings are treated as unrated on rebuild."""
    import csv as _csv

    from fantasy_football.export import write_tiers_csv

    _seed(session)
    ratings = {"prb0": 300.0, "prb1": 280.0,       # genuinely rated
               "prb2": 0.0, "prb3": -0.1, "prb4": -0.2}  # poisoned ladder
    out = tmp_path / "master.csv"
    write_tiers_csv(session, str(out), ratings=ratings, year=2025,
                    config=LeagueConfig(teams=1))
    rows = {r["key"]: r for r in _csv.DictReader(out.open())}
    # Ladder players got value-scale ratings back, ranked by production.
    assert float(rows["prb2"]["rating"]) > 0.5
    assert float(rows["prb2"]["rating"]) > float(rows["prb3"]["rating"])
    # The genuinely rated pair is untouched.
    assert float(rows["prb0"]["rating"]) == 300.0


def test_tier_notes_round_trip(session, tmp_path):
    """Notes written per tier come back via the CLI reader (carry-forward path)."""
    _seed(session)
    from fantasy_football.cli import _read_tier_notes
    from fantasy_football.export import write_tiers_csv

    out = tmp_path / "master.csv"
    write_tiers_csv(session, str(out), ratings={"prb0": 500.0, "prb1": 120.0},
                    notes={("RB", 1): "Proven elite vets"}, year=2025,
                    config=LeagueConfig(teams=1))
    notes = _read_tier_notes(str(out))
    assert notes[("RB", 1)] == "Proven elite vets"


def test_read_depth_overrides(tmp_path):
    from fantasy_football.cli import _read_depth_overrides

    p = tmp_path / "depth_overrides.csv"
    p.write_text("player,backup\nBijan Robinson,Tyler Allgeier\n,missing\nX,\n")
    over = _read_depth_overrides(str(p))
    assert over == {"bijan robinson": "Tyler Allgeier"}
    assert _read_depth_overrides(str(tmp_path / "nope.csv")) == {}


def test_webapp_payload_carries_packet_data(session, tmp_path):
    """data.js embeds prices, backups, the teams table, and the top-200 table."""
    import json

    from fantasy_football.export import write_webapp_data

    _seed(session)
    out = tmp_path / "data.js"
    write_webapp_data(
        session, str(out), year=2025, config=LeagueConfig(teams=1),
        prices={"prb0": 57.0},
        backups={"rb0": ("rb5", "RB5")},
        starters={("GB", "QB"): ["Some QB"], ("GB", "WR"): ["W One", "W Two"]},
        backup_overrides={"rb1": "Hand Picked"},
    )
    payload = json.loads(out.read_text().split("window.FF_DATA = ", 1)[1].rstrip(";\n"))

    rbs = {e["name"]: e for e in payload["positions"]["RB"]}
    assert rbs["RB0"]["price"] == 57.0
    assert rbs["RB0"]["bkp"] == "RB5"            # depth chart
    assert rbs["RB1"]["bkp"] == "Hand Picked"    # manual override wins
    assert rbs["RB2"]["bkp"] == "RB3"            # heuristic: next same-team RB

    gb = next(t for t in payload["teams"] if t["team"] == "GB")
    assert gb["pf"] == 20 and gb["qb"] == "Some QB" and gb["wr2"] == "W Two"

    assert payload["top200"][0][0] == "RB0"      # best scorer first
    assert len(payload["top200_headers"]) == len(payload["top200"][0])


def test_load_coaching_new_role_flags(session, tmp_path, monkeypatch):
    """Explicit hc_new/oc_new columns are honored; without hc_new, HC newness
    is auto-detected against last season's schedule coaches."""
    from fantasy_football.ingest import nflverse as nv

    _seed(session)

    # Explicit flags win.
    p = tmp_path / "coaching.csv"
    p.write_text("team,head_coach,offensive_coordinator,hc_new,oc_new\n"
                 "GB,Matt LaFleur,Adam Stenavich,,1\n"
                 "CHI,Ben Johnson,Press Taylor,1,\n")
    nv.load_coaching(session, str(p))
    teams = {t.abbreviation: t for t in session.query(Team).all()}
    assert not teams["GB"].hc_new and teams["GB"].oc_new
    assert teams["CHI"].hc_new and not teams["CHI"].oc_new

    # No hc_new column -> compare against last season's coaches (mocked).
    monkeypatch.setattr(nv, "latest_head_coaches",
                        lambda: {"GB": "Matt LaFleur", "CHI": "Matt Eberflus"})
    p2 = tmp_path / "coaching2.csv"
    p2.write_text("team,head_coach,offensive_coordinator\n"
                  "GB,Matt LaFleur,Adam Stenavich\n"
                  "CHI,Ben Johnson,Press Taylor\n")
    nv.load_coaching(session, str(p2))
    teams = {t.abbreviation: t for t in session.query(Team).all()}
    assert not teams["GB"].hc_new       # same coach as last season
    assert teams["CHI"].hc_new          # changed -> flagged


def test_webapp_payload_carries_coach_flags(session, tmp_path):
    import json

    from fantasy_football.export import write_webapp_data

    _seed(session)
    gb = session.query(Team).filter_by(abbreviation="GB").one()
    gb.head_coach, gb.oc_new = "Someone New", True
    gb.offensive_coordinator = "Fresh OC"
    session.commit()
    out = tmp_path / "data.js"
    write_webapp_data(session, str(out), year=2025, config=LeagueConfig(teams=1))
    payload = json.loads(out.read_text().split("window.FF_DATA = ", 1)[1].rstrip(";\n"))
    gbt = next(t for t in payload["teams"] if t["team"] == "GB")
    assert gbt["ocN"] == 1 and gbt["hcN"] == 0


def test_confidence_blend_scales_with_comps(session, tmp_path):
    """Few picks nudge a player off his anchor; many picks dominate."""
    import csv as _csv

    from fantasy_football.export import write_tiers_csv

    _seed(session)
    base = {"prb0": 300.0, "prb1": 200.0, "prb2": 200.0}
    # Users rated rb1 and rb2 identically higher (260), but rb1 was compared
    # 60 times and rb2 only twice.
    write_tiers_csv(session, str(tmp_path / "m.csv"),
                    ratings=base,
                    user_ratings={"prb1": 260.0, "prb2": 260.0},
                    comps={"prb1": 60, "prb2": 2}, confidence=6,
                    year=2025, config=LeagueConfig(teams=1))
    rows = {r["key"]: float(r["rating"]) for r in _csv.DictReader((tmp_path / "m.csv").open())}
    assert rows["prb0"] == 300.0                       # untouched anchor
    assert 250 < rows["prb1"] <= 260                   # 60 comps: ~91% of the way
    assert 210 < rows["prb2"] < 230                    # 2 comps: ~25% of the way
    assert rows["prb1"] > rows["prb2"]

    # Legacy pick file (no comps info) keeps the full-override behavior.
    write_tiers_csv(session, str(tmp_path / "m2.csv"),
                    ratings=base, user_ratings={"prb1": 260.0}, comps={},
                    year=2025, config=LeagueConfig(teams=1))
    rows2 = {r["key"]: float(r["rating"]) for r in _csv.DictReader((tmp_path / "m2.csv").open())}
    assert rows2["prb1"] == 260.0


def test_read_comps_hardened(tmp_path):
    from fantasy_football.cli import _merge_comps, _read_comps

    p = tmp_path / "picks.csv"
    p.write_text("key,rating,comps\nprb0,250,12\nprb1,240,\nbad key,1,5\nprb2,230,-3\n")
    comps = _read_comps(str(p))
    assert comps == {"prb0": 12, "prb2": 0}   # blank skipped, bad key skipped, clamped
    q = tmp_path / "picks2.csv"
    q.write_text("key,rating,comps\nprb0,260,8\n")
    assert _merge_comps([str(p), str(q)]) == {"prb0": 20, "prb2": 0}
    r = tmp_path / "legacy.csv"
    r.write_text("key,rating\nprb0,260\n")
    assert _read_comps(str(r)) == {}


def test_safe_cell_guards_formulas():
    from fantasy_football.export import _safe_cell

    assert _safe_cell("=1+1") == "'=1+1"
    assert _safe_cell("+danger") == "'+danger"
    assert _safe_cell("@x") == "'@x"
    assert _safe_cell("-2") == "'-2"
    assert _safe_cell("Ja'Marr Chase") == "Ja'Marr Chase"   # normal text untouched
    assert _safe_cell(123) == 123
